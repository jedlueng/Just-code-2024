import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Define the weekly schedule with time slots
week_schedule = {
    'Time': ['9:00 - 9:30 AM', '9:30 - 10:00 AM', '10:00 - 10:10 AM', '10:10 - 10:20 AM', '10:20 - 10:30 AM', 
             '10:30 - 11:30 AM', '11:30 AM - 12:00 PM', '1:00 - 2:00 PM', '2:00 - 5:00 PM', '5:00 - 6:00 PM', 
             '6:00 - 7:00 PM', '7:00 - 8:00 PM', '8:00 - 10:00 PM', '10:00 - 11:00 PM', '11:00 - 12:00 PM', 
             '12:00 - 9:00 AM'],
    'Monday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Work', 'Meal Prep', 
               'Boxing Class', 'Flexible', 'Dog Walk', 'Yoga Session', 'Dinner', 'Gaming/Movies', 
               'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep'],
    'Tuesday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Work', 'Meal Prep', 
                'Tennis Class', 'Work/Rest, socialize dog', 'Dog Walk', 'Yoga Session', 'Dinner', 'Gaming/Movies', 
                'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep'],
    'Wednesday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Work', 'Meal Prep', 
                  'Flexible', 'Mental Therapy', 'Dog Walk', 'Strength Training', 'Dinner', 'Gaming/Movies', 
                  'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep'],
    'Thursday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Work', 'Meal Prep', 
                 'Tennis Class', 'Work/Rest, socialize dog', 'Dog Walk', 'Yoga Session', 'Dinner', 'Gaming/Movies', 
                 'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep'],
    'Friday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Work', 'Meal Prep', 
               'Dog Swimming', 'Flexible', 'Dog Walk', 'Strength Training', 'Dinner', 'Gaming/Movies', 
               'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep'],
    'Saturday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Flexible', 'Meal Prep', 
                 'Mall Walk & Cafe', 'Meet Family', 'Dog Walk', 'Strength Training', 'Dinner', 'Gaming/Movies', 
                 'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep'],
    'Sunday': ['Wake up', 'Breakfast', 'Nootropics', 'Cold Water', 'Play with Beagle', 'Flexible', 'Meal Prep', 
               'Mall Walk & Cafe', 'Meet Family', 'Dog Walk', 'Rest/Light Activities', 'Dinner', 'Gaming/Movies', 
               'Meditate & Stretch', 'Talking with Girlfriend', 'Sleep']
}

# Convert the dictionary to a DataFrame
df = pd.DataFrame(week_schedule)

# Save the DataFrame to an Excel file
excel_path = 'weekly_schedule.xlsx'
df.to_excel(excel_path, index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(excel_path)
ws = wb.active

# Apply the headers
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(bottom=Side(style="thin"))

# Style the rows
alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
default_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
    fill = alt_fill if r_idx % 2 == 0 else default_fill
    for cell in row:
        cell.fill = fill

# Save the styled workbook
wb.save(excel_path)

print("Spreadsheet created and styled successfully.")
