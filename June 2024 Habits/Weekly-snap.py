import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the weekly schedule with 1-hour time slots
week_schedule = {
    'Time': ['9:00 - 10:00 AM', '10:00 - 11:00 AM', '11:00 AM - 12:00 PM', '12:00 - 1:00 PM', 
             '1:00 - 2:00 PM', '2:00 - 3:00 PM', '3:00 - 4:00 PM', '4:00 - 5:00 PM', '5:00 - 6:00 PM', 
             '6:00 - 7:00 PM', '7:00 - 8:00 PM', '8:00 - 9:00 PM', '9:00 - 10:00 PM', '10:00 - 11:00 PM', 
             '11:00 PM - 12:00 AM', '12:00 - 9:00 AM'],
    'Monday': ['🌅 Wake up', '🍳 Breakfast', '💻 Work', '🥗 Meal Prep', '🥊 Boxing Class', '🕊️ Flexible', '🕊️ Flexible', '🕊️ Flexible', 
               '🐕 Dog Walk', '🧘 Yoga Session', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
               '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep'],
    'Tuesday': ['🌅 Wake up', '🍳 Breakfast', '💻 Work', '🥗 Meal Prep', '🎾 Tennis Class', '🐕 Work/Rest, socialize dog', '🕊️ Flexible', '🕊️ Flexible', 
                '🐕 Dog Walk', '🧘 Yoga Session', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
                '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep'],
    'Wednesday': ['🌅 Wake up', '🍳 Breakfast', '💻 Work', '🥗 Meal Prep', '🕊️ Flexible', '🧠 Mental Therapy', '🕊️ Flexible', '🕊️ Flexible', 
                  '🐕 Dog Walk', '🏋️ Strength Training', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
                  '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep'],
    'Thursday': ['🌅 Wake up', '🍳 Breakfast', '💻 Work', '🥗 Meal Prep', '🎾 Tennis Class', '🐕 Work/Rest, socialize dog', '🕊️ Flexible', '🕊️ Flexible', 
                 '🐕 Dog Walk', '🧘 Yoga Session', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
                 '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep'],
    'Friday': ['🌅 Wake up', '🍳 Breakfast', '💻 Work', '🥗 Meal Prep', '🏊 Dog Swimming', '🕊️ Flexible', '🕊️ Flexible', '🕊️ Flexible', 
               '🐕 Dog Walk', '🏋️ Strength Training', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
               '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep'],
    'Saturday': ['🌅 Wake up', '🍳 Breakfast', '🕊️ Flexible', '🥗 Meal Prep', '🏬 Mall Walk & Cafe', '👪 Meet Family', '🕊️ Flexible', '🕊️ Flexible', 
                 '🐕 Dog Walk', '🏋️ Strength Training', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
                 '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep'],
    'Sunday': ['🌅 Wake up', '🍳 Breakfast', '🕊️ Flexible', '🥗 Meal Prep', '🏬 Mall Walk & Cafe', '👪 Meet Family', '🕊️ Flexible', '🕊️ Flexible', 
               '🐕 Dog Walk', '🧘 Rest/Light Activities', '🍽️ Dinner', '🎮 Gaming/Movies', '🎮 Gaming/Movies', 
               '🧘 Meditate & Stretch', '💬 Talking with Girlfriend', '💤 Sleep']
}

# Convert the dictionary to a DataFrame
df = pd.DataFrame(week_schedule)

# Define color mapping for each activity
color_mapping = {
    '🌅 Wake up': 'FFFF00', '🍳 Breakfast': 'FFA07A', '💻 Work': '00BFFF', '🥗 Meal Prep': '98FB98', '🥊 Boxing Class': 'FF4500',
    '🕊️ Flexible': 'D3D3D3', '🐕 Dog Walk': 'ADFF2F', '🧘 Yoga Session': 'AFEEEE', '🍽️ Dinner': 'FFA500', '🎮 Gaming/Movies': 'FF6347',
    '🧘 Meditate & Stretch': 'D8BFD8', '💬 Talking with Girlfriend': 'B0E0E6', '💤 Sleep': 'FFC0CB', '🎾 Tennis Class': 'FFD700',
    '🐕 Work/Rest, socialize dog': 'F0E68C', '🧠 Mental Therapy': 'CD5C5C', '🏋️ Strength Training': '708090', '🏊 Dog Swimming': 'FF69B4',
    '🏬 Mall Walk & Cafe': 'F4A460', '👪 Meet Family': '8B4513', '🧘 Rest/Light Activities': 'DAA520'
}

# Create an Excel writer object and write the DataFrame to an Excel file
excel_file = 'weekly_habit_snapshot_with_emoji.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Schedule')
    workbook = writer.book
    worksheet = writer.sheets['Schedule']

    # Apply formatting to the worksheet
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.alignment = alignment
            if cell.value in color_mapping:
                cell.fill = PatternFill(start_color=color_mapping[cell.value], end_color=color_mapping[cell.value], fill_type='solid')

    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

print(f"Excel file saved as '{excel_file}'")
