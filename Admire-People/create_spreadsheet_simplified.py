import subprocess
import sys

# Ensure the required library is installed:
try:
    import emoji
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "emoji"])
    import emoji

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Admired People"

# Define column headers
headers = [
    "Name", "Occupation", "Platform", "Specialization/Field", "Reasons for Admiration",
    "Trustworthiness", "Rating", "NAS Location", "Notes", "Country/Region"
]
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Add headers to the worksheet
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Function to generate star ratings
def generate_stars(rating):
    return emoji.emojize(':star:') * rating

# Sample data
data = [
    ["John Doe", "Fitness Trainer", "YouTube", "Fitness", f"{emoji.emojize(':muscle:')} Expertise in strength training", f"{emoji.emojize(':thumbs_up:')} High", generate_stars(5), "NAS/AdmiredPeople/JohnDoe", "Provides well-researched content and evidence-based tips.", "USA"],
    ["Jane Smith", "Tech Blogger", "Blog", "Technology", f"{emoji.emojize(':bulb:')} Integrity and honesty", f"{emoji.emojize(':thumbs_up:')} Medium", generate_stars(4), "NAS/AdmiredPeople/JaneSmith", "Occasionally makes minor errors but corrects them quickly.", "UK"],
    ["Mark Johnson", "Entrepreneur", "Twitter", "Business", f"{emoji.emojize(':chart_increasing:')} Consistency in advice", f"{emoji.emojize(':thumbs_up:')} High", generate_stars(5), "NAS/AdmiredPeople/MarkJohnson", "Successful track record and transparent about failures.", "Canada"]
]

# Add data to the worksheet
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        if col_num == 6:  # Trustworthiness column
            if "High" in cell_value:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif "Medium" in cell_value:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Adjust column widths
for col_num in range(1, len(headers) + 1):
    column_letter = get_column_letter(col_num)
    ws.column_dimensions[column_letter].width = 20

# Save the workbook
wb.save("Admired_People_Simplified.xlsx")
print("Spreadsheet created successfully!")
